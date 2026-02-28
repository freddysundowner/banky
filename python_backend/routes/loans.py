from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import List, Optional
from decimal import Decimal
from datetime import datetime, timedelta
from models.database import get_db
from models.master import Organization
from models.tenant import LoanApplication, LoanProduct, Member, LoanGuarantor, LoanExtraCharge, Transaction, Staff, LoanInstalment, CollateralItem, CollateralType, CollateralInsurance
from schemas.tenant import LoanApplicationCreate, LoanApplicationUpdate, LoanApplicationResponse, LoanApplicationAction, LoanDisbursement, LoanGuarantorCreate
from routes.auth import get_current_user
from middleware.demo_guard import require_not_demo
from routes.common import get_tenant_session_context, require_permission, require_any_permission, require_role
from services.code_generator import generate_txn_code

def try_send_sms(tenant_session, template_type: str, phone: str, name: str, context: dict, member_id=None, loan_id=None):
    """Try to send SMS, fail silently if SMS not configured"""
    try:
        from routes.sms import send_sms_with_template
        if phone:
            send_sms_with_template(tenant_session, template_type, phone, name, context, member_id=member_id, loan_id=loan_id)
    except Exception as e:
        print(f"[SMS] Failed to send {template_type}: {e}")

def post_disbursement_to_gl(tenant_session, loan, member, disbursement_method: str, interest_deducted_upfront: bool):
    """Post loan disbursement to the General Ledger"""
    try:
        from accounting.service import AccountingService, post_loan_disbursement
        
        svc = AccountingService(tenant_session)
        svc.seed_default_accounts()
        
        member_name = f"{member.first_name} {member.last_name}"
        
        print(f"[GL] Posting loan disbursement {loan.application_number}: amount={loan.amount}, net={loan.amount_disbursed}, interest_upfront={interest_deducted_upfront}")
        
        je = post_loan_disbursement(
            svc,
            member_id=str(member.id),
            loan_id=str(loan.id),
            principal_amount=loan.amount,
            net_disbursed=loan.amount_disbursed,
            interest_amount=loan.total_interest if interest_deducted_upfront else Decimal("0"),
            processing_fee=loan.processing_fee or Decimal("0"),
            insurance_fee=loan.insurance_fee or Decimal("0"),
            disbursement_method=disbursement_method,
            description=f"Loan disbursement - {member_name} - {loan.application_number}",
            appraisal_fee=loan.appraisal_fee or Decimal("0"),
            excise_duty=loan.excise_duty or Decimal("0")
        )
        print(f"[GL] Posted disbursement to GL: {loan.application_number}, journal entry: {je.entry_number if je else 'None'}")
    except Exception as e:
        import traceback
        print(f"[GL] Failed to post disbursement {loan.application_number} to GL: {e}")
        traceback.print_exc()

router = APIRouter()

def get_periodic_rate(interest_rate: Decimal, interest_rate_period: str, repayment_frequency: str) -> Decimal:
    periods_per_year = {"daily": 365, "weekly": 52, "bi_weekly": 26, "monthly": 12}
    rate_as_decimal = interest_rate / Decimal("100")
    
    freq_periods = periods_per_year.get(repayment_frequency, 12)
    rate_periods = periods_per_year.get(interest_rate_period, 12)
    
    if interest_rate_period == "annual":
        return rate_as_decimal / Decimal(str(freq_periods))
    elif rate_periods == freq_periods:
        return rate_as_decimal
    else:
        annual_rate = rate_as_decimal * Decimal(str(rate_periods))
        return annual_rate / Decimal(str(freq_periods))

def term_months_to_instalments(term_months: int, repayment_frequency: str) -> int:
    periods_per_year = {"daily": 365, "weekly": 52, "bi_weekly": 26, "monthly": 12}
    ppy = periods_per_year.get(repayment_frequency, 12)
    return max(round(term_months * ppy / 12), 1)

def instalments_to_term_months(num_instalments: int, repayment_frequency: str) -> int:
    periods_per_year = {"daily": 365, "weekly": 52, "bi_weekly": 26, "monthly": 12}
    ppy = periods_per_year.get(repayment_frequency, 12)
    return max(round(num_instalments * 12 / ppy), 1)

def calculate_loan(amount: Decimal, term_months: int, interest_rate: Decimal, interest_type: str = "reducing_balance", repayment_frequency: str = "monthly", interest_rate_period: str = "monthly"):
    periodic_rate = get_periodic_rate(interest_rate, interest_rate_period, repayment_frequency)
    num_instalments = term_months_to_instalments(term_months, repayment_frequency)

    n = num_instalments
    if interest_type == "flat":
        total_interest = amount * periodic_rate * n
        total_repayment = amount + total_interest
        periodic_payment = total_repayment / n if n > 0 else Decimal("0")
    else:
        if periodic_rate > 0:
            periodic_payment = amount * (periodic_rate * (1 + periodic_rate) ** n) / ((1 + periodic_rate) ** n - 1)
        else:
            periodic_payment = amount / n if n > 0 else Decimal("0")
        total_repayment = periodic_payment * n
        total_interest = total_repayment - amount

    return {
        "total_interest": round(total_interest, 2),
        "total_repayment": round(total_repayment, 2),
        "monthly_repayment": round(periodic_payment, 2),
        "num_instalments": n
    }

def generate_code(db: Session, prefix: str):
    count = db.query(func.count(LoanApplication.id)).scalar() or 0
    return f"{prefix}{count + 1:04d}"

@router.get("/{org_id}/loans")
async def list_loans(org_id: str, status: str = None, statuses: str = None, member_id: str = None, branch_id: str = None, page: int = 1, page_size: int = 20, search: str = None, product_id: str = None, date_from: str = None, date_to: str = None, user=Depends(get_current_user), db: Session = Depends(get_db)):
    from routes.common import get_branch_filter
    
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_permission(membership, "loans:read", db)
    tenant_session = tenant_ctx.create_session()
    try:
        query = tenant_session.query(LoanApplication)
        
        staff_branch_id = get_branch_filter(user)
        
        if staff_branch_id:
            branch_member_ids = [m.id for m in tenant_session.query(Member).filter(Member.branch_id == staff_branch_id).all()]
            query = query.filter(LoanApplication.member_id.in_(branch_member_ids))
        elif branch_id:
            branch_member_ids = [m.id for m in tenant_session.query(Member).filter(Member.branch_id == branch_id).all()]
            query = query.filter(LoanApplication.member_id.in_(branch_member_ids))
        
        if statuses:
            status_list = [s.strip() for s in statuses.split(",") if s.strip()]
            query = query.filter(LoanApplication.status.in_(status_list))
        elif status:
            query = query.filter(LoanApplication.status == status)
        if member_id:
            query = query.filter(LoanApplication.member_id == member_id)
        if product_id:
            query = query.filter(LoanApplication.loan_product_id == product_id)
        if date_from:
            try:
                from_date = datetime.strptime(date_from, "%Y-%m-%d")
                query = query.filter(LoanApplication.created_at >= from_date)
            except ValueError:
                pass
        if date_to:
            try:
                to_date = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
                query = query.filter(LoanApplication.created_at < to_date)
            except ValueError:
                pass
        
        if search:
            search_term = f"%{search.strip()}%"
            matching_member_ids = [m.id for m in tenant_session.query(Member).filter(
                or_(
                    Member.first_name.ilike(search_term),
                    Member.last_name.ilike(search_term),
                    Member.member_number.ilike(search_term),
                )
            ).all()]
            query = query.filter(
                or_(
                    LoanApplication.application_number.ilike(search_term),
                    LoanApplication.member_id.in_(matching_member_ids) if matching_member_ids else False,
                )
            )
        
        total = query.count()
        
        if page < 1:
            page = 1
        if page_size < 1:
            page_size = 20
        if page_size > 100:
            page_size = 100
        
        loans = query.order_by(LoanApplication.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
        
        member_ids = list(set(l.member_id for l in loans if l.member_id))
        product_ids = list(set(l.loan_product_id for l in loans if l.loan_product_id))
        staff_ids = list(set(
            sid for l in loans for sid in [l.created_by_id, l.reviewed_by_id] if sid
        ))
        
        members_map = {}
        if member_ids:
            for m in tenant_session.query(Member).filter(Member.id.in_(member_ids)).all():
                members_map[m.id] = m
        
        products_map = {}
        if product_ids:
            for p in tenant_session.query(LoanProduct).filter(LoanProduct.id.in_(product_ids)).all():
                products_map[p.id] = p
        
        staff_map = {}
        if staff_ids:
            for s in tenant_session.query(Staff).filter(Staff.id.in_(staff_ids)).all():
                staff_map[s.id] = s
        
        result = []
        for loan in loans:
            loan_dict = LoanApplicationResponse.model_validate(loan).model_dump()
            member = members_map.get(loan.member_id)
            product = products_map.get(loan.loan_product_id)
            created_by = staff_map.get(loan.created_by_id) if loan.created_by_id else None
            reviewed_by = staff_map.get(loan.reviewed_by_id) if loan.reviewed_by_id else None
            
            loan_dict["member_name"] = f"{member.first_name} {member.last_name}" if member else ""
            loan_dict["member_first_name"] = member.first_name if member else ""
            loan_dict["member_last_name"] = member.last_name if member else ""
            loan_dict["member_id_number"] = member.id_number if member else ""
            loan_dict["product_name"] = product.name if product else ""
            loan_dict["repayment_frequency"] = getattr(product, 'repayment_frequency', 'monthly') if product else "monthly"
            loan_dict["created_by_name"] = f"{created_by.first_name} {created_by.last_name}" if created_by else None
            loan_dict["reviewed_by_name"] = f"{reviewed_by.first_name} {reviewed_by.last_name}" if reviewed_by else None
            result.append(loan_dict)
        
        total_pages = (total + page_size - 1) // page_size
        
        return {
            "data": result,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": total_pages,
            }
        }
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.post("/{org_id}/loans")
async def create_loan(org_id: str, data: LoanApplicationCreate, user=Depends(get_current_user), db: Session = Depends(get_db)):
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_permission(membership, "loans:write", db)
    tenant_session = tenant_ctx.create_session()
    try:
        member = tenant_session.query(Member).filter(Member.id == data.member_id).first()
        if not member:
            raise HTTPException(status_code=404, detail="Member not found")
        
        if not member.is_active:
            raise HTTPException(status_code=400, detail="Cannot apply for loan: Member is not active")
        
        product = tenant_session.query(LoanProduct).filter(LoanProduct.id == data.loan_product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Loan product not found")
        
        active_loan_statuses = ["pending", "approved", "disbursed", "defaulted", "restructured"]
        
        if not getattr(product, 'allow_multiple_loans', True):
            existing_loan = tenant_session.query(LoanApplication).filter(
                LoanApplication.member_id == data.member_id,
                LoanApplication.loan_product_id == data.loan_product_id,
                LoanApplication.status.in_(active_loan_statuses)
            ).first()
            if existing_loan:
                raise HTTPException(
                    status_code=400,
                    detail=f"Member already has an active {product.name} loan ({existing_loan.application_number}). This product does not allow multiple loans."
                )
        
        if getattr(product, 'require_good_standing', False):
            active_loans = tenant_session.query(LoanApplication).filter(
                LoanApplication.member_id == data.member_id,
                LoanApplication.status.in_(["disbursed", "defaulted", "restructured"])
            ).all()
            for active_loan in active_loans:
                overdue_instalments = tenant_session.query(LoanInstalment).filter(
                    LoanInstalment.loan_id == active_loan.id,
                    LoanInstalment.status == "overdue"
                ).count()
                if overdue_instalments > 0:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Member has overdue payments on loan {active_loan.application_number}. All existing loans must be in good standing before applying for this product."
                    )
        
        # Shares-based eligibility checks (warning only - enforcement at disbursement)
        member_shares = Decimal(str(member.shares_balance or 0))
        shares_multiplier = Decimal(str(product.shares_multiplier or 0))
        min_shares_required = Decimal(str(product.min_shares_required or 0))
        
        eligibility_warning = None
        # Check minimum shares requirement (only if configured > 0)
        if min_shares_required > 0 and member_shares < min_shares_required:
            eligibility_warning = f"Warning: Member needs at least {min_shares_required:,.2f} in shares to qualify. Current shares: {member_shares:,.2f}. Disbursement will be blocked until this is met."
        # Check if loan amount exceeds shares-based limit (only if multiplier is configured > 0)
        elif shares_multiplier > 0:
            max_eligible_amount = member_shares * shares_multiplier
            if data.amount > max_eligible_amount:
                eligibility_warning = f"Warning: Loan amount exceeds eligible limit. Maximum ({member_shares:,.2f} × {shares_multiplier}x): {max_eligible_amount:,.2f}. Disbursement will be blocked until this is met."
        
        if data.amount < product.min_amount or data.amount > product.max_amount:
            raise HTTPException(status_code=400, detail=f"Amount must be between {product.min_amount} and {product.max_amount}")
        
        if data.term_months < product.min_term_months or data.term_months > product.max_term_months:
            raise HTTPException(status_code=400, detail=f"Term must be between {product.min_term_months} and {product.max_term_months} months")
        
        rate_period = getattr(product, 'interest_rate_period', None) or 'monthly'
        calc = calculate_loan(data.amount, data.term_months, product.interest_rate, product.interest_type, getattr(product, 'repayment_frequency', 'monthly'), rate_period)
        
        processing_fee = data.amount * (product.processing_fee or Decimal("0")) / Decimal("100")
        insurance_fee = data.amount * (product.insurance_fee or Decimal("0")) / Decimal("100")
        appraisal_fee = data.amount * (getattr(product, 'appraisal_fee', None) or Decimal("0")) / Decimal("100")
        
        base_fees = processing_fee + insurance_fee + appraisal_fee
        excise_duty_rate = getattr(product, 'excise_duty_rate', None) or Decimal("0")
        excise_duty = round(base_fees * excise_duty_rate / Decimal("100"), 2)
        total_fees = round(base_fees + excise_duty, 2)
        
        cli_rate = getattr(product, 'credit_life_insurance_rate', None) or Decimal("0")
        cli_freq = getattr(product, 'credit_life_insurance_freq', None) or "annual"
        
        effective_rate = get_periodic_rate(product.interest_rate, rate_period, getattr(product, 'repayment_frequency', 'monthly')) * Decimal("100")
        
        code = generate_code(tenant_session, "LN")
        
        current_staff = tenant_session.query(Staff).filter(Staff.email == user.email).first()
        
        loan = LoanApplication(
            application_number=code,
            member_id=data.member_id,
            loan_product_id=data.loan_product_id,
            amount=data.amount,
            term_months=data.term_months,
            interest_rate=round(effective_rate, 4),
            total_interest=calc["total_interest"],
            total_repayment=calc["total_repayment"],
            monthly_repayment=calc["monthly_repayment"],
            processing_fee=round(processing_fee, 2),
            insurance_fee=round(insurance_fee, 2),
            appraisal_fee=round(appraisal_fee, 2),
            excise_duty=excise_duty,
            total_fees=total_fees,
            credit_life_insurance_rate=cli_rate,
            credit_life_insurance_freq=cli_freq,
            purpose=data.purpose,
            disbursement_method=data.disbursement_method,
            disbursement_account=data.disbursement_account,
            disbursement_phone=data.disbursement_phone,
            status="pending",
            created_by_id=current_staff.id if current_staff else None
        )
        
        tenant_session.add(loan)
        tenant_session.commit()
        
        if data.guarantors and product.requires_guarantor:
            for g in data.guarantors:
                guarantor_member = tenant_session.query(Member).filter(Member.id == g.guarantor_id).first()
                if not guarantor_member:
                    continue
                
                guarantor = LoanGuarantor(
                    loan_id=loan.id,
                    guarantor_id=g.guarantor_id,
                    amount_guaranteed=g.amount_guaranteed,
                    status="pending"
                )
                tenant_session.add(guarantor)
            tenant_session.commit()

        if data.extra_charges:
            for ec in data.extra_charges:
                charge = LoanExtraCharge(
                    loan_id=loan.id,
                    charge_name=ec.charge_name,
                    amount=ec.amount
                )
                tenant_session.add(charge)
            tenant_session.commit()
        
        tenant_session.refresh(loan)
        response = LoanApplicationResponse.model_validate(loan).model_dump()
        if eligibility_warning:
            response["eligibility_warning"] = eligibility_warning
        return response
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.get("/{org_id}/loans/export")
async def export_loans(org_id: str, export_type: str = "all", status: str = None, product_id: str = None, branch_id: str = None, date_from: str = None, date_to: str = None, search: str = None, user=Depends(get_current_user), db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import date
    from routes.common import get_branch_filter
    
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_permission(membership, "loans:read", db)
    tenant_session = tenant_ctx.create_session()
    try:
        today = date.today()
        
        active_statuses = ["disbursed", "defaulted", "restructured"]
        unpaid_instalment_statuses = ["pending", "partial", "overdue"]
        
        if export_type == "due_today":
            due_loan_ids = [i.loan_id for i in tenant_session.query(LoanInstalment).filter(
                LoanInstalment.due_date == today,
                LoanInstalment.status.in_(unpaid_instalment_statuses)
            ).all()]
            query = tenant_session.query(LoanApplication).filter(
                LoanApplication.id.in_(due_loan_ids),
                LoanApplication.status.in_(active_statuses)
            )
            title = f"Loans Due Today - {today.strftime('%d %b %Y')}"
        elif export_type == "overdue":
            overdue_loan_ids = [i.loan_id for i in tenant_session.query(LoanInstalment).filter(
                LoanInstalment.due_date < today,
                LoanInstalment.status.in_(unpaid_instalment_statuses)
            ).all()]
            query = tenant_session.query(LoanApplication).filter(
                LoanApplication.id.in_(overdue_loan_ids),
                LoanApplication.status.in_(active_statuses)
            )
            title = f"Overdue Loans - {today.strftime('%d %b %Y')}"
        elif export_type == "due_this_week":
            week_end = today + timedelta(days=(6 - today.weekday()))
            due_loan_ids = [i.loan_id for i in tenant_session.query(LoanInstalment).filter(
                LoanInstalment.due_date >= today,
                LoanInstalment.due_date <= week_end,
                LoanInstalment.status.in_(unpaid_instalment_statuses)
            ).all()]
            query = tenant_session.query(LoanApplication).filter(
                LoanApplication.id.in_(due_loan_ids),
                LoanApplication.status.in_(active_statuses)
            )
            title = f"Loans Due This Week - {today.strftime('%d %b %Y')} to {week_end.strftime('%d %b %Y')}"
        elif export_type == "filtered":
            query = tenant_session.query(LoanApplication)
            if status:
                query = query.filter(LoanApplication.status == status)
            title = f"Filtered Loans Export - {today.strftime('%d %b %Y')}"
        else:
            query = tenant_session.query(LoanApplication)
            if status:
                query = query.filter(LoanApplication.status == status)
            else:
                query = query.filter(LoanApplication.status.in_(active_statuses))
            title = f"Loan Applications Export - {today.strftime('%d %b %Y')}"
        
        staff_branch_id = get_branch_filter(user)
        if staff_branch_id:
            branch_member_ids = [m.id for m in tenant_session.query(Member).filter(Member.branch_id == staff_branch_id).all()]
            query = query.filter(LoanApplication.member_id.in_(branch_member_ids))
        elif branch_id:
            branch_member_ids = [m.id for m in tenant_session.query(Member).filter(Member.branch_id == branch_id).all()]
            query = query.filter(LoanApplication.member_id.in_(branch_member_ids))
        
        if product_id:
            query = query.filter(LoanApplication.loan_product_id == product_id)
        if date_from:
            try:
                query = query.filter(LoanApplication.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
            except ValueError:
                pass
        if date_to:
            try:
                query = query.filter(LoanApplication.created_at < datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
            except ValueError:
                pass
        if search:
            search_term = f"%{search.strip()}%"
            matching_ids = [m.id for m in tenant_session.query(Member).filter(
                or_(Member.first_name.ilike(search_term), Member.last_name.ilike(search_term), Member.member_number.ilike(search_term))
            ).all()]
            query = query.filter(
                or_(
                    LoanApplication.application_number.ilike(search_term),
                    LoanApplication.member_id.in_(matching_ids) if matching_ids else False,
                )
            )
        
        loans = query.order_by(LoanApplication.created_at.desc()).all()
        
        member_ids = list(set(l.member_id for l in loans if l.member_id))
        product_ids_list = list(set(l.loan_product_id for l in loans if l.loan_product_id))
        members_map = {m.id: m for m in tenant_session.query(Member).filter(Member.id.in_(member_ids)).all()} if member_ids else {}
        products_map = {p.id: p for p in tenant_session.query(LoanProduct).filter(LoanProduct.id.in_(product_ids_list)).all()} if product_ids_list else {}
        
        loan_ids = [l.id for l in loans]
        instalments = tenant_session.query(LoanInstalment).filter(
            LoanInstalment.loan_id.in_(loan_ids)
        ).all() if loan_ids else []
        
        instalment_map = {}
        for inst in instalments:
            if inst.loan_id not in instalment_map:
                instalment_map[inst.loan_id] = []
            instalment_map[inst.loan_id].append(inst)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Loans"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        ws.merge_cells("A1:N1")
        title_cell = ws["A1"]
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color="1E3A5F")
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30
        
        ws.merge_cells("A2:N2")
        ws["A2"].value = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')} | Total Records: {len(loans)}"
        ws["A2"].font = Font(italic=True, color="666666", size=10)
        ws["A2"].alignment = Alignment(horizontal="center")
        
        headers = [
            "Loan #", "Member Name", "Member Number", "Phone", "Alt. Phone", "Email",
            "Product", "Amount Disbursed", "Outstanding Balance", "Amount Repaid",
            "Next Due Date", "Overdue Amount", "Days Overdue", "Status"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        overdue_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
        due_today_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
        
        for row_idx, loan in enumerate(loans, 5):
            member = members_map.get(loan.member_id)
            product = products_map.get(loan.loan_product_id)
            loan_instalments = instalment_map.get(loan.id, [])
            
            overdue_instalments = [i for i in loan_instalments if i.due_date < today and i.status in ("pending", "partial", "overdue")]
            overdue_amount = sum(
                float(i.expected_principal or 0) + float(i.expected_interest or 0) - float(i.paid_principal or 0) - float(i.paid_interest or 0)
                for i in overdue_instalments
            )
            days_overdue = 0
            if overdue_instalments:
                earliest_overdue = min(i.due_date for i in overdue_instalments)
                days_overdue = (today - earliest_overdue).days
            
            next_due = None
            unpaid_instalments = sorted(
                [i for i in loan_instalments if i.status in ("pending", "partial", "overdue")],
                key=lambda i: i.due_date
            )
            if unpaid_instalments:
                next_due = unpaid_instalments[0].due_date
            
            row_data = [
                loan.application_number,
                f"{member.first_name} {member.last_name}" if member else "",
                member.member_number if member else "",
                member.phone or "" if member else "",
                member.phone_secondary or "" if member else "",
                member.email or "" if member else "",
                product.name if product else "",
                float(loan.amount_disbursed or loan.amount or 0),
                float(loan.outstanding_balance or 0),
                float(loan.amount_repaid or 0),
                next_due.strftime("%d %b %Y") if next_due else "",
                round(overdue_amount, 2) if overdue_amount > 0 else "",
                days_overdue if days_overdue > 0 else "",
                loan.status.replace("_", " ").title(),
            ]
            
            is_overdue = days_overdue > 0
            is_due_today = next_due == today if next_due else False
            
            currency_cols = {8, 9, 10, 12}
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx in currency_cols and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                if is_overdue:
                    cell.fill = overdue_fill
                elif is_due_today:
                    cell.fill = due_today_fill
        
        col_widths = [14, 22, 16, 18, 16, 26, 24, 18, 18, 16, 16, 16, 14, 14]
        for idx, width in enumerate(col_widths, 1):
            col_letter = ws.cell(row=4, column=idx).column_letter
            ws.column_dimensions[col_letter].width = width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        safe_title = export_type.replace(" ", "_")
        filename = f"loans_{safe_title}_{today.strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.post("/{org_id}/loans/eligibility-check")
async def check_loan_eligibility(
    org_id: str,
    payload: dict,
    user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Quick field eligibility check — returns pass/fail for each criterion."""
    from datetime import date as date_type
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_any_permission(membership, ["loans:read", "loans:write"], db)
    tenant_session = tenant_ctx.create_session()
    try:
        member_id = payload.get("member_id")  # optional
        product_id = payload.get("loan_product_id")
        amount = Decimal(str(payload.get("amount", 0)))
        term_months = int(payload.get("term_months", 12))
        collateral_value = Decimal(str(payload.get("collateral_value", 0)))
        # Manual financial inputs (used when no member account exists)
        manual_savings = Decimal(str(payload.get("manual_savings", 0)))
        manual_shares = Decimal(str(payload.get("manual_shares", 0)))
        manual_deposits = Decimal(str(payload.get("manual_deposits", 0)))
        prospect_name = payload.get("prospect_name", "Prospect")

        # Load member from DB if provided
        member = None
        if member_id:
            member = tenant_session.query(Member).filter(Member.id == member_id).first()
            if not member:
                raise HTTPException(status_code=404, detail="Member not found")

        product = tenant_session.query(LoanProduct).filter(LoanProduct.id == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Loan product not found")

        # Resolve balances — from DB if member exists, otherwise from manual inputs
        if member:
            eff_savings = member.savings_balance or Decimal("0")
            eff_shares = member.shares_balance or Decimal("0")
            eff_deposits = member.deposits_balance or Decimal("0")
        else:
            eff_savings = manual_savings
            eff_shares = manual_shares
            eff_deposits = manual_deposits

        checks = []
        all_passed = True

        # 1. Member active (only when a member is linked)
        if member:
            is_active = member.status == "active"
            checks.append({
                "key": "member_status",
                "label": "Member Status",
                "passed": is_active,
                "message": "Member is active and in good standing." if is_active else f"Member status is '{member.status}'. Only active members can apply.",
            })
            if not is_active:
                all_passed = False
        else:
            checks.append({
                "key": "member_status",
                "label": "Member Status",
                "passed": True,
                "message": "Prospect check — membership status not verified. Client will need to be registered as an active member before applying.",
                "informational": True,
            })

        # 2. Amount within product range
        in_range = product.min_amount <= amount <= product.max_amount
        checks.append({
            "key": "amount_range",
            "label": "Loan Amount Range",
            "passed": in_range,
            "message": f"Requested {float(amount):,.2f} is within the product range of {float(product.min_amount):,.2f} – {float(product.max_amount):,.2f}." if in_range
                       else f"Requested {float(amount):,.2f} is outside the product range ({float(product.min_amount):,.2f} – {float(product.max_amount):,.2f}).",
        })
        if not in_range:
            all_passed = False

        # 3. Term within range
        term_ok = product.min_term_months <= term_months <= product.max_term_months
        checks.append({
            "key": "term_range",
            "label": "Loan Term",
            "passed": term_ok,
            "message": f"Requested term of {term_months} months is within the allowed range ({product.min_term_months}–{product.max_term_months} months)." if term_ok
                       else f"Requested term of {term_months} months is outside the allowed range ({product.min_term_months}–{product.max_term_months} months).",
        })
        if not term_ok:
            all_passed = False

        # 4. Shares-based eligibility
        shares_balance = eff_shares
        shares_multiplier = product.shares_multiplier or Decimal("0")
        min_shares_req = product.min_shares_required or Decimal("0")
        max_from_shares = shares_balance * shares_multiplier if shares_multiplier > 0 else None

        if shares_multiplier > 0:
            shares_ok = amount <= max_from_shares
            checks.append({
                "key": "shares_coverage",
                "label": "Shares Coverage",
                "passed": shares_ok,
                "message": f"Shares balance {float(shares_balance):,.2f} × {float(shares_multiplier)} = {float(max_from_shares):,.2f} covers the loan." if shares_ok
                           else f"Shares balance {float(shares_balance):,.2f} × {float(shares_multiplier)} = {float(max_from_shares):,.2f} — insufficient to cover {float(amount):,.2f}.",
            })
            if not shares_ok:
                all_passed = False
        elif min_shares_req > 0:
            shares_ok = shares_balance >= min_shares_req
            checks.append({
                "key": "min_shares",
                "label": "Minimum Shares",
                "passed": shares_ok,
                "message": f"Shares balance {float(shares_balance):,.2f} meets the minimum requirement of {float(min_shares_req):,.2f}." if shares_ok
                           else f"Shares balance {float(shares_balance):,.2f} is below the minimum required {float(min_shares_req):,.2f}.",
            })
            if not shares_ok:
                all_passed = False

        # 5. Savings balance info (informational)
        savings_total = eff_savings + eff_deposits
        checks.append({
            "key": "savings_info",
            "label": "Savings & Deposits",
            "passed": True,
            "message": f"Client has {float(savings_total):,.2f} in savings and deposits (shares: {float(shares_balance):,.2f}).",
            "informational": True,
        })

        # 6. Multiple loans check (only for existing members)
        active_loan_statuses = ["pending", "approved", "disbursed", "defaulted", "restructured"]
        if member and not product.allow_multiple_loans:
            existing = tenant_session.query(LoanApplication).filter(
                LoanApplication.member_id == member_id,
                LoanApplication.loan_product_id == product_id,
                LoanApplication.status.in_(active_loan_statuses)
            ).first()
            no_conflict = existing is None
            checks.append({
                "key": "multiple_loans",
                "label": "Active Loans",
                "passed": no_conflict,
                "message": "No existing active loan for this product." if no_conflict
                           else f"Member already has an active {product.name} loan ({existing.application_number}). This product does not allow multiple loans.",
            })
            if not no_conflict:
                all_passed = False
        elif not member and not product.allow_multiple_loans:
            checks.append({
                "key": "multiple_loans",
                "label": "Active Loans",
                "passed": True,
                "message": "Prospect check — cannot verify existing loans without a member account. Will be checked on application.",
                "informational": True,
            })

        # 7. Good standing check (only for existing members)
        if member and product.require_good_standing:
            active_loans = tenant_session.query(LoanApplication).filter(
                LoanApplication.member_id == member_id,
                LoanApplication.status.in_(["disbursed", "defaulted", "restructured"])
            ).all()
            today = date_type.today()
            overdue = False
            for loan in active_loans:
                overdue_inst = tenant_session.query(LoanInstalment).filter(
                    LoanInstalment.loan_id == loan.id,
                    LoanInstalment.status.in_(["pending", "partially_paid"]),
                    LoanInstalment.due_date < today
                ).first()
                if overdue_inst:
                    overdue = True
                    break
            checks.append({
                "key": "good_standing",
                "label": "Good Standing",
                "passed": not overdue,
                "message": "Member has no overdue loan repayments." if not overdue else "Member has overdue repayments on an existing loan. All loans must be in good standing.",
            })
            if overdue:
                all_passed = False
        elif not member and product.require_good_standing:
            checks.append({
                "key": "good_standing",
                "label": "Good Standing",
                "passed": True,
                "message": "Prospect check — good standing cannot be verified without a member account. Will be checked on application.",
                "informational": True,
            })

        # 8. Collateral check
        if product.requires_collateral and float(product.min_ltv_coverage or 0) > 0:
            min_coverage_pct = float(product.min_ltv_coverage) / 100.0
            required_collateral_value = float(amount) * min_coverage_pct
            provided = float(collateral_value)
            collateral_ok = provided >= required_collateral_value
            checks.append({
                "key": "collateral",
                "label": "Collateral Coverage",
                "passed": collateral_ok,
                "message": f"Provided collateral value {provided:,.2f} covers the required {required_collateral_value:,.2f} ({float(product.min_ltv_coverage):.0f}% of loan)." if collateral_ok
                           else f"Collateral value {provided:,.2f} is below the required {required_collateral_value:,.2f} ({float(product.min_ltv_coverage):.0f}% of loan amount).",
            })
            if not collateral_ok:
                all_passed = False
        elif product.requires_collateral and float(collateral_value) == 0:
            checks.append({
                "key": "collateral",
                "label": "Collateral Required",
                "passed": False,
                "message": "This product requires collateral. Please provide the estimated collateral value.",
            })
            all_passed = False

        # Calculate estimated monthly payment (reducing balance)
        monthly_rate = float(product.interest_rate) / 100.0
        if product.interest_rate_period == "annual":
            monthly_rate = monthly_rate / 12.0
        estimated_payment = 0.0
        if monthly_rate > 0 and term_months > 0:
            estimated_payment = float(amount) * (monthly_rate * (1 + monthly_rate) ** term_months) / ((1 + monthly_rate) ** term_months - 1)
        elif term_months > 0:
            estimated_payment = float(amount) / term_months

        # Fee estimates
        processing_fee = float(amount) * float(product.processing_fee or 0) / 100.0
        insurance_fee = float(amount) * float(product.insurance_fee or 0) / 100.0
        appraisal_fee = float(product.appraisal_fee or 0)
        total_fees = processing_fee + insurance_fee + appraisal_fee

        # Max eligible from shares
        max_eligible = None
        if shares_multiplier > 0:
            max_eligible = float(shares_balance * shares_multiplier)

        return {
            "eligible": all_passed,
            "is_prospect": member is None,
            "checks": checks,
            "member": {
                "id": member.id if member else None,
                "name": f"{member.first_name} {member.last_name}" if member else prospect_name,
                "member_number": member.member_number if member else None,
                "status": member.status if member else "prospect",
                "savings_balance": float(eff_savings),
                "shares_balance": float(eff_shares),
                "deposits_balance": float(eff_deposits),
            },
            "product": {
                "id": product.id,
                "name": product.name,
                "interest_rate": float(product.interest_rate),
                "interest_rate_period": product.interest_rate_period,
                "requires_guarantor": product.requires_guarantor,
                "min_guarantors": product.min_guarantors,
                "requires_collateral": product.requires_collateral,
            },
            "requested": {
                "amount": float(amount),
                "term_months": term_months,
                "collateral_value": float(collateral_value),
            },
            "estimates": {
                "monthly_payment": round(estimated_payment, 2),
                "processing_fee": round(processing_fee, 2),
                "insurance_fee": round(insurance_fee, 2),
                "appraisal_fee": round(appraisal_fee, 2),
                "total_fees": round(total_fees, 2),
                "total_repayable": round(estimated_payment * term_months, 2),
            },
            "max_eligible_amount": max_eligible,
        }
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.get("/{org_id}/loans/{loan_id}")
async def get_loan(org_id: str, loan_id: str, user=Depends(get_current_user), db: Session = Depends(get_db)):
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_permission(membership, "loans:read", db)
    tenant_session = tenant_ctx.create_session()
    try:
        loan = tenant_session.query(LoanApplication).filter(LoanApplication.id == loan_id).first()
        if not loan:
            raise HTTPException(status_code=404, detail="Loan not found")
        
        loan_dict = LoanApplicationResponse.model_validate(loan).model_dump()
        
        # Add member and product names
        member = tenant_session.query(Member).filter(Member.id == loan.member_id).first()
        product = tenant_session.query(LoanProduct).filter(LoanProduct.id == loan.loan_product_id).first()
        created_by = tenant_session.query(Staff).filter(Staff.id == loan.created_by_id).first() if loan.created_by_id else None
        reviewed_by = tenant_session.query(Staff).filter(Staff.id == loan.reviewed_by_id).first() if loan.reviewed_by_id else None
        
        loan_dict["member_name"] = f"{member.first_name} {member.last_name}" if member else ""
        loan_dict["member_first_name"] = member.first_name if member else ""
        loan_dict["member_last_name"] = member.last_name if member else ""
        loan_dict["product_name"] = product.name if product else ""
        loan_dict["product_requires_collateral"] = bool(product.requires_collateral) if product else False
        loan_dict["product_min_ltv_coverage"] = float(product.min_ltv_coverage or 0) if product else 0.0
        loan_dict["created_by_name"] = f"{created_by.first_name} {created_by.last_name}" if created_by else None
        loan_dict["reviewed_by_name"] = f"{reviewed_by.first_name} {reviewed_by.last_name}" if reviewed_by else None
        
        return loan_dict
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.put("/{org_id}/loans/{loan_id}/action")
async def process_loan_action(org_id: str, loan_id: str, data: LoanApplicationAction, user=Depends(get_current_user), db: Session = Depends(get_db)):
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    if data.action == "approve":
        require_permission(membership, "loans:approve")
    elif data.action == "reject":
        require_permission(membership, "loans:reject")
    else:
        require_permission(membership, "loans:process")
    tenant_session = tenant_ctx.create_session()
    try:
        loan = tenant_session.query(LoanApplication).filter(LoanApplication.id == loan_id).first()
        if not loan:
            raise HTTPException(status_code=404, detail="Loan not found")
        
        # Get current staff for tracking who reviewed the loan
        current_staff = tenant_session.query(Staff).filter(Staff.email == user.email).first()
        
        if data.action == "approve":
            if loan.status != "pending":
                raise HTTPException(status_code=400, detail="Loan is not pending")
            
            # Check guarantor requirements
            product = tenant_session.query(LoanProduct).filter(LoanProduct.id == loan.loan_product_id).first()
            if product and product.requires_guarantor:
                accepted_guarantors = tenant_session.query(LoanGuarantor).filter(
                    LoanGuarantor.loan_id == loan.id,
                    LoanGuarantor.status == "accepted"
                ).count()
                if accepted_guarantors == 0:
                    raise HTTPException(
                        status_code=400, 
                        detail="This loan requires at least one accepted guarantor before approval. Please add and get guarantors to accept first."
                    )

            # Check collateral requirements
            if product and product.requires_collateral:
                collateral_items = (
                    tenant_session.query(CollateralItem)
                    .join(CollateralType, CollateralItem.collateral_type_id == CollateralType.id)
                    .filter(CollateralItem.loan_id == loan.id)
                    .all()
                )
                if not collateral_items:
                    raise HTTPException(
                        status_code=400,
                        detail="This loan product requires collateral. Please register at least one collateral item before approving."
                    )
                min_coverage = float(product.min_ltv_coverage or 0)
                if min_coverage > 0:
                    total_ltv_value = sum(
                        float(item.appraised_value or item.declared_value or 0)
                        * float(item.collateral_type.ltv_percent or 0) / 100
                        for item in collateral_items
                    )
                    required_value = float(loan.amount) * min_coverage / 100
                    if total_ltv_value < required_value:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Collateral coverage is insufficient. Required LTV-adjusted value: KES {required_value:,.2f} ({min_coverage}% of loan). Current: KES {total_ltv_value:,.2f}. Please add more collateral or get existing items appraised."
                        )

            loan.status = "approved"
            loan.approved_at = datetime.utcnow()
            loan.reviewed_by_id = current_staff.id if current_staff else None
            
            # Send SMS notification for loan approval
            member = tenant_session.query(Member).filter(Member.id == loan.member_id).first()
            if member and member.phone:
                try_send_sms(
                    tenant_session, 
                    "loan_approved", 
                    member.phone, 
                    f"{member.first_name} {member.last_name}",
                    {"name": member.first_name, "amount": str(loan.amount)},
                    member_id=member.id,
                    loan_id=loan.id
                )
        
        elif data.action == "reject":
            if loan.status not in ["pending", "under_review"]:
                raise HTTPException(status_code=400, detail="Loan cannot be rejected in this status")
            loan.status = "rejected"
            loan.rejected_at = datetime.utcnow()
            loan.rejection_reason = data.reason
            loan.reviewed_by_id = current_staff.id if current_staff else None
            
            # Send SMS notification for loan rejection
            member = tenant_session.query(Member).filter(Member.id == loan.member_id).first()
            if member and member.phone:
                try_send_sms(
                    tenant_session, 
                    "loan_rejected", 
                    member.phone, 
                    f"{member.first_name} {member.last_name}",
                    {"name": member.first_name},
                    member_id=member.id,
                    loan_id=loan.id
                )
        
        elif data.action == "review":
            if loan.status != "pending":
                raise HTTPException(status_code=400, detail="Loan is not pending")
            loan.status = "under_review"
            loan.reviewed_by_id = current_staff.id if current_staff else None
        
        else:
            raise HTTPException(status_code=400, detail="Invalid action")
        
        tenant_session.commit()
        tenant_session.refresh(loan)
        
        # Return enriched response with staff names
        loan_dict = LoanApplicationResponse.model_validate(loan).model_dump()
        member = tenant_session.query(Member).filter(Member.id == loan.member_id).first()
        product = tenant_session.query(LoanProduct).filter(LoanProduct.id == loan.loan_product_id).first()
        created_by = tenant_session.query(Staff).filter(Staff.id == loan.created_by_id).first() if loan.created_by_id else None
        reviewed_by = tenant_session.query(Staff).filter(Staff.id == loan.reviewed_by_id).first() if loan.reviewed_by_id else None
        
        loan_dict["member_name"] = f"{member.first_name} {member.last_name}" if member else ""
        loan_dict["member_first_name"] = member.first_name if member else ""
        loan_dict["member_last_name"] = member.last_name if member else ""
        loan_dict["product_name"] = product.name if product else ""
        loan_dict["created_by_name"] = f"{created_by.first_name} {created_by.last_name}" if created_by else None
        loan_dict["reviewed_by_name"] = f"{reviewed_by.first_name} {reviewed_by.last_name}" if reviewed_by else None
        
        return loan_dict
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.post("/{org_id}/loans/{loan_id}/disburse")
async def disburse_loan(org_id: str, loan_id: str, data: LoanDisbursement, user=Depends(get_current_user), db: Session = Depends(get_db)):
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_permission(membership, "loans:process")
    tenant_session = tenant_ctx.create_session()
    try:
        loan = tenant_session.query(LoanApplication).filter(LoanApplication.id == loan_id).with_for_update().first()
        if not loan:
            raise HTTPException(status_code=404, detail="Loan not found")
        
        if loan.status != "approved":
            raise HTTPException(status_code=400, detail="Loan must be approved before disbursement")
        
        member = tenant_session.query(Member).filter(Member.id == loan.member_id).with_for_update().first()
        product = tenant_session.query(LoanProduct).filter(LoanProduct.id == loan.loan_product_id).first()
        
        if member and product:
            member_shares = Decimal(str(member.shares_balance or 0))
            shares_multiplier = Decimal(str(product.shares_multiplier or 0))
            min_shares_required = Decimal(str(product.min_shares_required or 0))
            
            # Check minimum shares requirement (only if configured > 0)
            if min_shares_required > 0 and member_shares < min_shares_required:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Cannot disburse: Member needs at least {min_shares_required:,.2f} in shares to qualify. Current shares: {member_shares:,.2f}"
                )
            
            # Check if loan amount exceeds shares-based limit (only if multiplier is configured > 0)
            if shares_multiplier > 0:
                max_eligible_amount = member_shares * shares_multiplier
                if loan.amount > max_eligible_amount:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Cannot disburse: Loan amount exceeds eligible limit. Maximum ({member_shares:,.2f} × {shares_multiplier}x): {max_eligible_amount:,.2f}"
                    )

            # Collateral check at disbursement (safety net)
            if product.requires_collateral:
                collateral_items = (
                    tenant_session.query(CollateralItem)
                    .join(CollateralType, CollateralItem.collateral_type_id == CollateralType.id)
                    .filter(CollateralItem.loan_id == loan.id)
                    .all()
                )
                if not collateral_items:
                    raise HTTPException(
                        status_code=400,
                        detail="Cannot disburse: This loan product requires collateral but none is registered against this loan."
                    )
                min_coverage = float(product.min_ltv_coverage or 0)
                if min_coverage > 0:
                    total_ltv_value = sum(
                        float(item.appraised_value or item.declared_value or 0)
                        * float(item.collateral_type.ltv_percent or 0) / 100
                        for item in collateral_items
                    )
                    required_value = float(loan.amount) * min_coverage / 100
                    if total_ltv_value < required_value:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Cannot disburse: Collateral LTV-adjusted value (KES {total_ltv_value:,.2f}) is below the required {min_coverage}% coverage (KES {required_value:,.2f})."
                        )

                # Insurance check: block disbursement if any collateral type requires
                # insurance but the item has no active, non-expired policy.
                today = datetime.utcnow().date()
                uninsured = []
                for item in collateral_items:
                    if not item.collateral_type.requires_insurance:
                        continue
                    active_policy = (
                        tenant_session.query(CollateralInsurance)
                        .filter(
                            CollateralInsurance.collateral_item_id == item.id,
                            CollateralInsurance.status == "active",
                            CollateralInsurance.expiry_date >= today,
                        )
                        .first()
                    )
                    if not active_policy:
                        uninsured.append(item.description)
                if uninsured:
                    items_list = ", ".join(f'"{d}"' for d in uninsured)
                    raise HTTPException(
                        status_code=400,
                        detail=f"Cannot disburse: The following collateral item(s) require an active insurance policy before disbursement: {items_list}."
                    )
        
        if data.disbursement_method == "mpesa":
            if not data.disbursement_phone:
                raise HTTPException(status_code=400, detail="Phone number required for M-Pesa disbursement")
            from middleware.demo_guard import is_demo_mode as _is_demo
            from models.tenant import OrganizationSettings as _OrgSettings
            mpesa_setting = tenant_session.query(_OrgSettings).filter(
                _OrgSettings.setting_key == "mpesa_enabled"
            ).first()
            mpesa_on = mpesa_setting and mpesa_setting.setting_value.lower() == "true" if mpesa_setting else False
            if not mpesa_on:
                raise HTTPException(status_code=400, detail="M-Pesa is not enabled for this organization")
            if not _is_demo():
                org = db.query(Organization).filter(Organization.id == org_id).first()
                if not org or (getattr(org, "currency", None) or "USD") != "KES":
                    raise HTTPException(status_code=400, detail="M-Pesa disbursement is only available for organizations using KES currency")
        
        if data.disbursement_method == "bank" and not data.disbursement_account:
            raise HTTPException(status_code=400, detail="Account number required for bank disbursement")
        
        net_amount = loan.amount - (loan.total_fees or Decimal("0"))
        
        # Check if interest should be deducted upfront
        deduct_interest_upfront = getattr(product, 'deduct_interest_upfront', False) if product else False
        
        if deduct_interest_upfront:
            # Deduct interest from disbursement amount
            net_amount = net_amount - (loan.total_interest or Decimal("0"))
            outstanding_balance = loan.amount
            freq = getattr(product, 'repayment_frequency', 'monthly') or 'monthly'
            n_inst = term_months_to_instalments(loan.term_months, freq)
            monthly_repayment = loan.amount / n_inst if n_inst > 0 else loan.amount
        else:
            outstanding_balance = loan.total_repayment
            monthly_repayment = loan.monthly_repayment
        
        loan.disbursement_method = data.disbursement_method
        loan.disbursement_account = data.disbursement_account
        loan.disbursement_phone = data.disbursement_phone
        loan.amount_disbursed = net_amount
        loan.outstanding_balance = outstanding_balance
        loan.interest_deducted_upfront = deduct_interest_upfront
        loan.monthly_repayment = monthly_repayment
        loan.amount_repaid = Decimal("0")
        freq_days = {"daily": 1, "weekly": 7, "bi_weekly": 14, "monthly": 30}
        product_freq = getattr(product, 'repayment_frequency', 'monthly') if product else 'monthly'
        loan.next_payment_date = (datetime.utcnow() + timedelta(days=freq_days.get(product_freq, 30))).date()
        
        is_mpesa = data.disbursement_method == "mpesa" and data.disbursement_phone
        
        loan.disbursed_at = datetime.utcnow()
        if is_mpesa:
            loan.status = "pending_disbursement"
        else:
            loan.status = "disbursed"
        
        txn_code = generate_txn_code()
        
        transaction = Transaction(
            transaction_number=txn_code,
            member_id=loan.member_id,
            transaction_type="loan_disbursement",
            account_type="loan",
            amount=net_amount,
            description=f"Loan disbursement for {loan.application_number}"
        )
        
        if is_mpesa:
            transaction.description = f"Loan disbursement (pending M-Pesa) for {loan.application_number}"
        
        tenant_session.add(transaction)
        
        if data.disbursement_method == "savings" and member:
            balance_before_disbursement = member.savings_balance or Decimal("0")
            member.savings_balance = balance_before_disbursement + net_amount
            savings_txn = Transaction(
                transaction_number=generate_txn_code(),
                member_id=loan.member_id,
                transaction_type="deposit",
                account_type="savings",
                amount=net_amount,
                balance_before=balance_before_disbursement,
                balance_after=member.savings_balance,
                description=f"Loan disbursement to savings - {loan.application_number}"
            )
            tenant_session.add(savings_txn)
        
        from services.instalment_service import generate_instalment_schedule
        generate_instalment_schedule(tenant_session, loan, product)
        
        if loan.total_insurance and loan.total_insurance > 0:
            loan.outstanding_balance = (loan.outstanding_balance or Decimal("0")) + loan.total_insurance
        
        tenant_session.commit()
        tenant_session.refresh(loan)
        
        if not is_mpesa and member:
            post_disbursement_to_gl(tenant_session, loan, member, data.disbursement_method, deduct_interest_upfront)
        
        mpesa_b2c_result = None
        if is_mpesa:
            try:
                phone = data.disbursement_phone.replace("+", "").replace(" ", "")
                if phone.startswith("0"):
                    phone = "254" + phone[1:]

                from routes.mpesa import initiate_b2c_disbursement
                mpesa_b2c_result = initiate_b2c_disbursement(
                    tenant_session, phone, net_amount,
                    remarks=f"Loan disbursement {loan.application_number}",
                    occasion=loan.application_number
                )
                if mpesa_b2c_result and mpesa_b2c_result.get("success"):
                    loan.status = "disbursed"
                    loan.disbursed_at = datetime.utcnow()
                    transaction.description = f"Loan disbursement for {loan.application_number}"
                    tenant_session.commit()
                    if member:
                        post_disbursement_to_gl(tenant_session, loan, member, data.disbursement_method, deduct_interest_upfront)
                    
                print(f"[M-Pesa B2C] Disbursement for {loan.application_number}: {mpesa_b2c_result}")
                
                if mpesa_b2c_result and not mpesa_b2c_result.get("success"):
                    loan.status = "approved"
                    loan.disbursed_at = None
                    loan.amount_disbursed = None
                    loan.outstanding_balance = None
                    loan.amount_repaid = None
                    loan.next_payment_date = None
                    from models.tenant import LoanInstalment
                    tenant_session.query(LoanInstalment).filter(LoanInstalment.loan_id == loan.id).delete()
                    tenant_session.delete(transaction)
                    tenant_session.commit()
                    print(f"[M-Pesa B2C] Reverted loan {loan.application_number} back to approved due to B2C failure")
                    
            except Exception as e:
                print(f"[M-Pesa B2C] Failed to send disbursement for {loan.application_number}: {e}")
                loan.status = "approved"
                loan.disbursed_at = None
                loan.amount_disbursed = None
                loan.outstanding_balance = None
                loan.amount_repaid = None
                loan.next_payment_date = None
                from models.tenant import LoanInstalment
                tenant_session.query(LoanInstalment).filter(LoanInstalment.loan_id == loan.id).delete()
                tenant_session.delete(transaction)
                tenant_session.commit()
                mpesa_b2c_result = {"success": False, "error": str(e)}

        if not is_mpesa and member and member.phone:
            try_send_sms(
                tenant_session,
                "loan_disbursed",
                member.phone,
                f"{member.first_name} {member.last_name}",
                {
                    "name": member.first_name,
                    "amount": str(net_amount),
                    "monthly_repayment": str(loan.monthly_repayment or 0),
                    "due_date": str(loan.next_payment_date) if loan.next_payment_date else "N/A"
                },
                member_id=member.id,
                loan_id=loan.id
            )
        
        # Return enriched response with staff names
        loan_dict = LoanApplicationResponse.model_validate(loan).model_dump()
        member = tenant_session.query(Member).filter(Member.id == loan.member_id).first()
        product = tenant_session.query(LoanProduct).filter(LoanProduct.id == loan.loan_product_id).first()
        created_by = tenant_session.query(Staff).filter(Staff.id == loan.created_by_id).first() if loan.created_by_id else None
        reviewed_by = tenant_session.query(Staff).filter(Staff.id == loan.reviewed_by_id).first() if loan.reviewed_by_id else None
        
        loan_dict["member_name"] = f"{member.first_name} {member.last_name}" if member else ""
        loan_dict["member_first_name"] = member.first_name if member else ""
        loan_dict["member_last_name"] = member.last_name if member else ""
        loan_dict["product_name"] = product.name if product else ""
        loan_dict["created_by_name"] = f"{created_by.first_name} {created_by.last_name}" if created_by else None
        loan_dict["reviewed_by_name"] = f"{reviewed_by.first_name} {reviewed_by.last_name}" if reviewed_by else None
        
        response = {
            "message": "Loan disbursed successfully",
            "loan": loan_dict,
            "disbursement": {
                "method": data.disbursement_method,
                "amount": float(net_amount),
                "processing_fee": float(loan.processing_fee or 0),
                "insurance_fee": float(loan.insurance_fee or 0),
                "appraisal_fee": float(loan.appraisal_fee or 0),
                "excise_duty": float(loan.excise_duty or 0),
                "total_fees": float(loan.total_fees or 0),
                "interest_deducted_upfront": deduct_interest_upfront,
                "interest_deducted": float(loan.total_interest or 0) if deduct_interest_upfront else 0
            }
        }
        if mpesa_b2c_result is not None:
            response["mpesa_b2c"] = mpesa_b2c_result
        return response
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.get("/{org_id}/loans/{loan_id}/summary")
async def get_loan_summary(org_id: str, loan_id: str, user=Depends(get_current_user), db: Session = Depends(get_db)):
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_permission(membership, "loans:read", db)
    tenant_session = tenant_ctx.create_session()
    try:
        loan = tenant_session.query(LoanApplication).filter(LoanApplication.id == loan_id).first()
        if not loan:
            raise HTTPException(status_code=404, detail="Loan not found")
        
        member = tenant_session.query(Member).filter(Member.id == loan.member_id).first()
        product = tenant_session.query(LoanProduct).filter(LoanProduct.id == loan.loan_product_id).first()
        guarantors = tenant_session.query(LoanGuarantor).filter(LoanGuarantor.loan_id == loan_id).all()
        
        interest_deducted_upfront = getattr(loan, 'interest_deducted_upfront', False)
        if interest_deducted_upfront:
            total_installments_due = float(loan.amount)
        else:
            total_installments_due = float(loan.total_repayment or 0)
        
        return {
            "loan": LoanApplicationResponse.model_validate(loan),
            "member": {
                "id": member.id,
                "member_number": member.member_number,
                "name": f"{member.first_name} {member.last_name}",
                "phone": member.phone,
                "savings": float(member.savings_balance or 0)
            } if member else None,
            "product": {
                "id": product.id,
                "name": product.name,
                "interest_rate": float(product.interest_rate),
                "interest_type": product.interest_type
            } if product else None,
            "guarantors": [
                {
                    "id": g.id,
                    "guarantor_id": g.guarantor_id,
                    "amount_guaranteed": float(g.amount_guaranteed),
                    "status": g.status
                } for g in guarantors
            ],
            "calculations": {
                "principal": float(loan.amount),
                "interest": float(loan.total_interest or 0),
                "interest_deducted_upfront": interest_deducted_upfront,
                "fees": float(loan.total_fees or 0),
                "processing_fee": float(loan.processing_fee or 0),
                "insurance_fee": float(loan.insurance_fee or 0),
                "appraisal_fee": float(loan.appraisal_fee or 0),
                "excise_duty": float(loan.excise_duty or 0),
                "total_insurance": float(loan.total_insurance or 0),
                "total_due": total_installments_due,
                "amount_paid": float(loan.amount_repaid or 0),
                "outstanding": float(loan.outstanding_balance or 0)
            }
        }
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.get("/{org_id}/loans/{loan_id}/instalments")
async def get_loan_instalments(org_id: str, loan_id: str, user=Depends(get_current_user), db: Session = Depends(get_db)):
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_permission(membership, "loans:read", db)
    tenant_session = tenant_ctx.create_session()
    try:
        from models.tenant import LoanInstalment
        instalments = tenant_session.query(LoanInstalment).filter(
            LoanInstalment.loan_id == loan_id
        ).order_by(LoanInstalment.instalment_number).all()
        
        return [{
            "id": str(i.id),
            "instalment_number": i.instalment_number,
            "due_date": str(i.due_date),
            "expected_principal": float(i.expected_principal),
            "expected_interest": float(i.expected_interest),
            "expected_penalty": float(i.expected_penalty),
            "expected_insurance": float(getattr(i, 'expected_insurance', None) or 0),
            "paid_principal": float(i.paid_principal),
            "paid_interest": float(i.paid_interest),
            "paid_penalty": float(i.paid_penalty),
            "paid_insurance": float(getattr(i, 'paid_insurance', None) or 0),
            "total_due": float(i.expected_principal + i.expected_interest + i.expected_penalty + (getattr(i, 'expected_insurance', None) or 0)),
            "total_paid": float(i.paid_principal + i.paid_interest + i.paid_penalty + (getattr(i, 'paid_insurance', None) or 0)),
            "balance": float(
                (i.expected_principal + i.expected_interest + i.expected_penalty + (getattr(i, 'expected_insurance', None) or 0)) -
                (i.paid_principal + i.paid_interest + i.paid_penalty + (getattr(i, 'paid_insurance', None) or 0))
            ),
            "status": i.status,
            "paid_at": str(i.paid_at) if i.paid_at else None,
        } for i in instalments]
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.delete("/{org_id}/loans/{loan_id}", dependencies=[Depends(require_not_demo)])
async def delete_loan(org_id: str, loan_id: str, user=Depends(get_current_user), db: Session = Depends(get_db)):
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_role(membership, ["owner", "admin"])
    tenant_session = tenant_ctx.create_session()
    try:
        loan = tenant_session.query(LoanApplication).filter(LoanApplication.id == loan_id).first()
        if not loan:
            raise HTTPException(status_code=404, detail="Loan not found")
        
        if loan.status in ["disbursed", "paid"]:
            raise HTTPException(status_code=400, detail="Cannot delete disbursed or paid loans")
        
        tenant_session.query(LoanGuarantor).filter(LoanGuarantor.loan_id == loan_id).delete()
        tenant_session.delete(loan)
        tenant_session.commit()
        return {"message": "Loan deleted"}
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.put("/{org_id}/loans/{loan_id}")
async def update_loan(org_id: str, loan_id: str, data: LoanApplicationUpdate, user=Depends(get_current_user), db: Session = Depends(get_db)):
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_permission(membership, "loans:write", db)
    tenant_session = tenant_ctx.create_session()
    try:
        loan = tenant_session.query(LoanApplication).filter(LoanApplication.id == loan_id).first()
        if not loan:
            raise HTTPException(status_code=404, detail="Loan not found")
        
        if loan.status not in ["pending", "under_review", "cancelled", "rejected"]:
            raise HTTPException(status_code=400, detail="Can only edit loans with status pending, under_review, cancelled, or rejected")
        
        if data.loan_product_id:
            product = tenant_session.query(LoanProduct).filter(LoanProduct.id == data.loan_product_id).first()
            if not product:
                raise HTTPException(status_code=404, detail="Loan product not found")
            loan.loan_product_id = data.loan_product_id
            edit_rate_period = getattr(product, 'interest_rate_period', None) or 'monthly'
            edit_effective_rate = get_periodic_rate(product.interest_rate, edit_rate_period, getattr(product, 'repayment_frequency', 'monthly')) * Decimal("100")
            loan.interest_rate = round(edit_effective_rate, 4)
        else:
            product = tenant_session.query(LoanProduct).filter(LoanProduct.id == loan.loan_product_id).first()
        
        if data.amount is not None:
            if data.amount < product.min_amount or data.amount > product.max_amount:
                raise HTTPException(status_code=400, detail=f"Amount must be between {product.min_amount} and {product.max_amount}")
            loan.amount = data.amount
        
        if data.term_months is not None:
            if data.term_months < product.min_term_months or data.term_months > product.max_term_months:
                raise HTTPException(status_code=400, detail=f"Term must be between {product.min_term_months} and {product.max_term_months} months")
            loan.term_months = data.term_months
        
        if data.purpose is not None:
            loan.purpose = data.purpose
        if data.disbursement_method is not None:
            loan.disbursement_method = data.disbursement_method
        if data.disbursement_account is not None:
            loan.disbursement_account = data.disbursement_account
        if data.disbursement_phone is not None:
            loan.disbursement_phone = data.disbursement_phone
        
        edit_rp = getattr(product, 'interest_rate_period', None) or 'monthly'
        calc = calculate_loan(loan.amount, loan.term_months, product.interest_rate, product.interest_type, getattr(product, 'repayment_frequency', 'monthly'), edit_rp)
        loan.total_interest = calc["total_interest"]
        loan.total_repayment = calc["total_repayment"]
        loan.monthly_repayment = calc["monthly_repayment"]
        
        processing_fee = loan.amount * (product.processing_fee or Decimal("0")) / Decimal("100")
        insurance_fee = loan.amount * (product.insurance_fee or Decimal("0")) / Decimal("100")
        appraisal_fee = loan.amount * (getattr(product, 'appraisal_fee', None) or Decimal("0")) / Decimal("100")
        loan.processing_fee = round(processing_fee, 2)
        loan.insurance_fee = round(insurance_fee, 2)
        loan.appraisal_fee = round(appraisal_fee, 2)
        
        base_fees = processing_fee + insurance_fee + appraisal_fee
        excise_duty_rate = getattr(product, 'excise_duty_rate', None) or Decimal("0")
        loan.excise_duty = round(base_fees * excise_duty_rate / Decimal("100"), 2)
        loan.total_fees = round(base_fees + loan.excise_duty, 2)
        
        loan.credit_life_insurance_rate = getattr(product, 'credit_life_insurance_rate', None) or Decimal("0")
        loan.credit_life_insurance_freq = getattr(product, 'credit_life_insurance_freq', None) or "annual"
        
        if loan.status in ["under_review", "cancelled", "rejected"]:
            loan.status = "pending"
            loan.rejection_reason = None
            loan.rejected_at = None
        
        tenant_session.commit()
        tenant_session.refresh(loan)
        return LoanApplicationResponse.model_validate(loan)
    finally:
        tenant_session.close()
        tenant_ctx.close()

@router.put("/{org_id}/loans/{loan_id}/cancel")
async def cancel_loan(org_id: str, loan_id: str, user=Depends(get_current_user), db: Session = Depends(get_db)):
    tenant_ctx, membership = get_tenant_session_context(org_id, user, db)
    require_permission(membership, "loans:write", db)
    tenant_session = tenant_ctx.create_session()
    try:
        loan = tenant_session.query(LoanApplication).filter(LoanApplication.id == loan_id).first()
        if not loan:
            raise HTTPException(status_code=404, detail="Loan not found")
        
        if loan.status not in ["pending", "under_review"]:
            raise HTTPException(status_code=400, detail="Can only cancel loans with status pending or under_review")
        
        loan.status = "cancelled"
        tenant_session.commit()
        tenant_session.refresh(loan)
        return LoanApplicationResponse.model_validate(loan)
    finally:
        tenant_session.close()
        tenant_ctx.close()

# Alias routes for frontend compatibility - loan-applications maps to loans
@router.get("/{org_id}/loan-applications")
async def list_loan_applications(org_id: str, status: str = None, statuses: str = None, member_id: str = None, branch_id: str = None, page: int = 1, page_size: int = 20, search: str = None, product_id: str = None, date_from: str = None, date_to: str = None, user=Depends(get_current_user), db: Session = Depends(get_db)):
    return await list_loans(org_id, status=status, statuses=statuses, member_id=member_id, branch_id=branch_id, page=page, page_size=page_size, search=search, product_id=product_id, date_from=date_from, date_to=date_to, user=user, db=db)

@router.post("/{org_id}/loan-applications")
async def create_loan_application(org_id: str, data: LoanApplicationCreate, user=Depends(get_current_user), db: Session = Depends(get_db)):
    return await create_loan(org_id, data, user, db)

@router.get("/{org_id}/loan-applications/{loan_id}")
async def get_loan_application(org_id: str, loan_id: str, user=Depends(get_current_user), db: Session = Depends(get_db)):
    return await get_loan(org_id, loan_id, user, db)

@router.post("/{org_id}/loan-applications/{loan_id}/approve")
async def approve_loan_application(org_id: str, loan_id: str, user=Depends(get_current_user), db: Session = Depends(get_db)):
    return await process_loan_action(org_id, loan_id, LoanApplicationAction(action="approve"), user, db)

@router.post("/{org_id}/loan-applications/{loan_id}/reject")
async def reject_loan_application(org_id: str, loan_id: str, data: dict = None, user=Depends(get_current_user), db: Session = Depends(get_db)):
    reason = data.get("reason", "") if data else ""
    return await process_loan_action(org_id, loan_id, LoanApplicationAction(action="reject", reason=reason), user, db)

@router.post("/{org_id}/loan-applications/{loan_id}/disburse")
async def disburse_loan_application(org_id: str, loan_id: str, data: Optional[LoanDisbursement] = None, user=Depends(get_current_user), db: Session = Depends(get_db)):
    if data is None:
        data = LoanDisbursement(disbursement_method="cash")
    return await disburse_loan(org_id, loan_id, data, user, db)

@router.put("/{org_id}/loan-applications/{loan_id}")
async def update_loan_application(org_id: str, loan_id: str, data: LoanApplicationUpdate, user=Depends(get_current_user), db: Session = Depends(get_db)):
    return await update_loan(org_id, loan_id, data, user, db)

@router.put("/{org_id}/loan-applications/{loan_id}/cancel")
async def cancel_loan_application(org_id: str, loan_id: str, user=Depends(get_current_user), db: Session = Depends(get_db)):
    return await cancel_loan(org_id, loan_id, user, db)
